import json
import openpyxl

from openpyxl.utils import get_column_letter


def read_category_data():
    with open("Category storage.json", newline='') as data_file:
        return json.load(data_file)


def write_category_data(data):
    with open("Category storage.json", 'w') as data_file:
        return json.dump(data, data_file)


def read_goods_data():
    with open("item_storage.json", newline='') as data_file:
        return json.load(data_file)


def write_goods_data(data):
    with open("item_storage.json", 'w') as data_file:
        return json.dump(data, data_file)


def max_item_id():
    with open("item_storage.json", newline='') as data_file:
        reader = json.load(data_file)
        if reader:
            return max(int(item['item_id']) for item in reader)


def clear_goods_data():
    with open("item_storage.json", 'w') as data_file:
        return json.dump([], data_file)


def read_order_data():
    with open("order_storage.json", newline='') as data_file:
        return json.load(data_file)


def write_order_data(data):
    with open("order_storage.json", 'w') as data_file:
        return json.dump(data, data_file)


def max_order_id():
    with open("order_storage.json", newline='') as data_file:
        reader = json.load(data_file)
        if reader:
            return max(int(item['order_id']) for item in reader)


def write_to_excel(sh1, sh2, sh3, sh4, min_date, max_date):
    lst_of_sheets = []
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'sales_by_category'
    ws1.append(['Category', 'Quantity_sold', 'Revenue_received'])
    for item in sh1:
        ws1.append(item)
    lst_of_sheets.append(ws1)

    wb.create_sheet('items_sold')
    ws2 = wb['items_sold']
    ws2.append(['item_id', 'name', 'quantity_sold'])
    for item in sh2:
        item.pop()
        ws2.append(item)
    lst_of_sheets.append(ws2)

    wb.create_sheet('orders')
    ws3 = wb['orders']
    ws3.append(['order_id', 'amount', 'created_at'])
    for item in sh3:
        ws3.append(item)
    lst_of_sheets.append(ws3)

    wb.create_sheet('total_metrics')
    ws4 = wb['total_metrics']
    ws4.append(['Total_revenue', sh4[0][0]])
    ws4.append(['Items_in_orders', sh4[0][1]])
    ws4.append(['Most_popular_category', sh4[1]])
    ws4.append(['Most_popular_item', sh4[2]])

    lst_of_sheets.append(ws4)

    for sheet in lst_of_sheets:
        for i in range(1, sheet.max_column + 1):
            column_letter = get_column_letter(i)
            max_length = 0
            for cell in sheet[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as err:
                    print(err)
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

    wb.save(f'statistics_period_{min_date}-{max_date}.xlsx')
    wb.close()
